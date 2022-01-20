from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

from .onto.leavedonto import OntoManager, OntTrie, LeavedOnto


def onto_from_tagged(in_file, out_file, resources, basis_onto=None):
    # load words and tags
    tagged = get_entries(in_file)

    # load all available ontos
    om = OntoManager(basis_onto)
    om.batch_merge_to_onto(resources.values())

    # populate new onto:
    trie = tagged_to_trie(tagged, om.onto1, in_file.stem.split("_")[0])
    onto = LeavedOnto(trie, out_file)
    onto.convert2yaml()


def get_entries(in_file):
    wb = load_workbook(in_file)
    ws = wb.active

    # from sheet to list of lists
    tagged = []
    max_row, max_col = coordinate_to_tuple(ws.dimensions.split(":")[1])
    for r in range(1, max_row + 1, 4):
        for col in range(1, max_col + 1):
            # ignoring the first column containing the numbers
            word = ws.cell(r, col).value
            pos = ws.cell(r + 1, col).value
            level = ws.cell(r + 2, col).value
            entry = (word, pos, level)
            if pos and level and entry not in tagged:
                tagged.append(entry)

    return tagged


def tagged_to_trie(tagged, onto_basis, origin):
    trie = OntTrie()
    trie.legend = onto_basis.ont.legend
    for word, pos, level in tagged:
        found = onto_basis.ont.find_entries(prefix=pos, lemma=word)
        if found:
            for path, entries in found:
                for e in entries:
                    found_level = onto_basis.get_field_value(e, "level")
                    if found_level == level:
                        onto_basis.set_field_value(e, "origin", origin)
                        trie.add(path, e)
        else:
            path = [pos, "to_organize"]
            parts = {"word": word, "POS": pos, "level": level}
            entry = [parts[l] if l in parts else "" for l in onto_basis.ont.legend]
            onto_basis.set_field_value(entry, "origin", origin)
            trie.add(path, entry)
    return trie
