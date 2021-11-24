from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment


def generate_xlsx(in_file, out_file):
    font = 'Jomolhari'
    ft_sent = Font(font, size=12)
    alignmnt = Alignment(horizontal="left", vertical="center")

    wb = Workbook()
    wb.remove(wb.get_sheet_by_name('Sheet'))

    lines = in_file.read_text().split('\n')
    for s, sent in enumerate(lines):
        ws = wb.create_sheet(title=str(s))
        words = sent.split(' ')
        for i in range(2, 12):
            for n, w in enumerate(words):
                ws.cell(row=i, column=n+1).value = w
                ws.cell(row=i, column=n + 1).font = ft_sent
                ws.cell(row=i, column=n + 1).alignment = alignmnt

    wb.save(out_file)
