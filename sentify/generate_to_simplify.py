from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .onto.leavedonto import OntoManager
from .dataval import DataVal
from .utils import resize_sheet


def generate_to_simplify(in_file, out_file, resources, l_colors, basis_onto=None):
    how_many_copied_lines = 10
    font = "Jomolhari"
    ft_sent = Font(font, size=12)
    alignmnt = Alignment(horizontal="left", vertical="center")

    wb = Workbook()
    wb.remove(wb.get_sheet_by_name("Sheet"))

    # add sentences to sheets
    # extract triples from _totag.xlsx: sentence, pos and levels
    sentences = extract_tagged(in_file)  # in_file.read_text().lstrip("\ufeff").split("\n")

    # prepare the synonyms
    om = OntoManager(basis_onto)
    om.batch_merge_to_onto(list(resources.values()))
    onto = om.onto1

    # prepare synonym data validation
    dv = DataVal(wb)

    val_num = 0
    for n, sentence in enumerate(sentences):
        sheet_name = str(n)
        ws = wb.create_sheet(title=sheet_name)
        for i in range(2, how_many_copied_lines + 2):
            for m, parts in enumerate(sentence):
                cell = ws.cell(row=i, column=m + 1)
                word, pos, level = parts
                cell.value = word
                cell.font = ft_sent
                cell.alignment = alignmnt

                # only adding synonyms and level color code to first row
                if i > 2:
                    continue

                cell.fill = PatternFill("solid", fgColor=l_colors[level])

                # ## SYNONYMS ## #
                # extract all synonyms
                found = onto.find_word(word)
                syns = [word]
                for path, entries in found:
                    for e in entries:
                        lemma = onto.get_field_value(e, "lemma")
                        if lemma:
                            syns.append(lemma)
                        syn = onto.get_field_value(e, "synonyms")
                        if syn:
                            syns.extend([a for a in syn.split(" ") if a])
                syns = list(set(syns))

                # pass if no synonyms
                if (len(syns) == 1 and syns[0] != word) or len(syns) > 1:
                    # add synonyms as data validation to cell
                    val_name = f"{val_num} {word}"
                    dv.add_validator(val_name, syns)
                    dv.add_val_to_cell(val_name, sheet_name, idx=cell.coordinate)
                    val_num += 1

        resize_sheet(ws)

    wb.save(out_file)


def extract_tagged(in_file):
    lines_per_sentence = 4

    wb = load_workbook(in_file)
    dump = list(wb.active.values)
    sentences = []
    for i in range(0, len(dump), lines_per_sentence):
        words = list(dump[i])
        pos = list(dump[i+1])
        levels = list(dump[i+2])
        sent = []
        i = 0
        while i < len(words):
            if words[i] and pos[i] and levels[i]:
                sent.append((words[i], pos[i], levels[i]))
            else:
                break
            i += 1
        sentences.append(sent)

    return sentences
