from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection

from .onto.leavedonto import OntoManager
from .dataval import DataVal
from .utils import resize_sheet


def rows_from_lines(lines):
    row_size = 12

    words = []
    for l in lines:
        words.extend(l.split(" "))

    rows = []
    cur_row = []
    while words:
        cur_row.append(words.pop(0))

        if len(cur_row) == row_size:
            rows.append(cur_row)
            cur_row = []
    if cur_row:
        rows.append(cur_row)

    return rows


def generate_to_tag(in_file, out_file, resources, l_colors=None):
    font = "Jomolhari"
    ft_words = Font(font, size=17, color="000c1d91")
    ft_pos = Font(font, size=13, color="004e4f54")
    ft_level = Font(size=11, color="004e4f54")
    new_bgcolor = PatternFill("solid", fgColor="0090f0a9")
    alignmnt = Alignment(horizontal="left", vertical="center")

    wb = Workbook()
    wb.remove(wb.get_sheet_by_name("Sheet"))

    # read input file into rows
    lines = in_file.read_text().lstrip("\ufeff").split("\n")
    rows = rows_from_lines(lines)

    # load available ontos
    main_onto = resources["general_onto"]
    other_ontos = [r for r in resources.values() if r.stem != "general_onto"]

    om = OntoManager(main_onto)
    om.batch_merge_to_onto(other_ontos)
    onto = om.onto1

    # prepare data validation for POS and levels
    dv = DataVal(wb)
    pos = [
        "ཚིག་ཕྲད།",
        "མིང་ཚིག",
        "བྱ་ཚིག",
        "རྒྱན་ཚིག",
        "བསྣན་ཚིག",
        "ཚབ་ཚིག",
        "ཚེག་ཤད།",
        "ཁྱད་ཚིག",
    ]
    dv.add_validator("POS", pos)
    levels = ["A0", "A1", "A2", "A2+", "B1", "B1+", "B2", "B2+", "C1", "C1+"]
    dv.add_validator("level", levels)

    # create sheet and fill it
    sheet_name = in_file.stem.split("_")[0]
    ws = wb.create_sheet(title=sheet_name)
    ws.protection.sheet = True
    for n, r in enumerate(rows):
        row = n * 4 + 1
        pos_row = row + 1
        level_row = pos_row + 1
        for m, el in enumerate(r):
            col = m + 1

            # check if word exists in onto
            found = onto.find_word(el)
            found_pos = found[0][0][0] if found else None
            entries = found[0][1] if found else None

            # add word to spreadsheet
            word_cell = ws.cell(row=row, column=col)
            word_cell.value = el
            word_cell.font = ft_words
            word_cell.alignment = alignmnt

            # add POS
            pos_cell = ws.cell(row=pos_row, column=col)
            pos_cell.protection = Protection(locked=False)
            pos_cell.value = found_pos if found_pos else ""
            pos_cell.font = ft_pos
            pos_cell.alignment = alignmnt
            dv.add_val_to_cell(
                val_name="POS", sheet_name=sheet_name, row=pos_row, col=col
            )
            if not entries:
                pos_cell.fill = new_bgcolor

            # add level
            level_cell = ws.cell(row=level_row, column=col)
            level_cell.protection = Protection(locked=False)
            level_cell.value = (
                onto.get_field_value(entries[0], "level") if entries else ""
            )
            level_cell.font = ft_level
            level_cell.alignment = alignmnt
            dv.add_val_to_cell(
                val_name="level", sheet_name=sheet_name, row=level_row, col=col
            )
            if not entries:
                level_cell.fill = new_bgcolor

        # set row height for each group of "word, POS and level"
        ws.row_dimensions[row].height = 30
        ws.row_dimensions[pos_row].height = 15
        ws.row_dimensions[level_row].height = 15
        ws.row_dimensions[level_row + 1].height = 30
        # size of empty row between two lines

    resize_sheet(ws, mode="width")
    wb.save(out_file)
