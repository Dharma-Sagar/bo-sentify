from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname


""" ***usage example***

wb = DataValWB()

# add sheets
# 1st el: adds validation to row if True
# 2nd el: row to add
data = [
    (True, ['Dog']),
    (True, ['Cat']),
    (False, ['empty', 'and', 'nothing'])
]
wb.add_sheet('Sheet', data)

# add data val sheet
values = ['a', 'b', 'c', 'd']
wb.add_val_sheet(values)

# save
wb.save('test.xlsx')
"""


class DataValWB:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.val_name = '_'
        self.dv = self.__set_dv()

    def add_sheet(self, sheet_name, rows):
        ws = self.wb.create_sheet(sheet_name)
        ws.add_data_validation(self.dv)
        for n, r in enumerate(rows):
            val, row = r
            for m, v in enumerate(row):
                cell = ws.cell(n+1, m+1)
                cell.value = v
                if val:
                    self.dv.add(cell)

    def add_val_sheet(self, row):
        ws = self.wb.create_sheet(self.val_name)
        for r in row:
            ws.append([r])

    def __set_dv(self):
        dv = DataValidation(type="list", formula1="{0}!$A$1:$A$10000".format(quote_sheetname(self.val_name)))
        # Optionally set a custom error message
        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'
        # Optionally set a custom prompt message
        dv.prompt = 'Please select from the list'
        dv.promptTitle = 'List Selection'
        return dv

    def save(self, name):
        self.wb.save(name)


wb = DataValWB()

# add sheets
# 1st el: adds validation to row if True
# 2nd el: row to add
data = [
    (True, ['Dog']),
    (True, ['Cat']),
    (False, ['empty', 'and', 'nothing'])
]
wb.add_sheet('Sheet', data)

# add data val sheet
values = ['a', 'b', 'c', 'd']
wb.add_val_sheet(values)

# save
wb.save('test.xlsx')